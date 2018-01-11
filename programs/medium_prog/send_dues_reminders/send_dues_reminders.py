# python send_dues_reminders.py 
# Рассылает сообщения на основании сведений из электронной 
# таблицы об уплате взносов.

# :::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::

import openpyxl, smtplib, sys
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

# :::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::

file = openpyxl.load_workbook('dues_records.xlsx')
sheet = file.active

last_col = sheet.max_column
latest_month = sheet.cell(row=1, column=last_col).value

# проверка состояние уплаты взносов для каждого члена клуба
unpaid_members = {}

for row in list(sheet.rows)[1:]:
	payment = row[last_col-1].value
	if payment != 'Оплачено':
		name = row[0].value
		email = row[1].value
		unpaid_members[name] = email



# установка учётных данных электроной почты
MAIL_SERVER = 'smtp.gmail.com'
MAIL_PORT = 465
MAIL_USERNAME = sys.argv[1]
MAIL_PASSWORD = sys.argv[2]


# отправка сообщений с напоминанием об уплате взносов
for user in unpaid_members:
	msg = MIMEMultipart('alternative')
	msg['Subject'] = "unpaid_members"
	msg['From'] = MAIL_USERNAME
	msg['To'] = unpaid_members[user]

	html = """
		<html>
			<head></head>
			<body>
				<p style="text-align:center;">Уважаемый <i style="color:green;">{}</i>.</p>
				<p style="text-align:center;padding-left:100px;padding-right:100px;"> 
					У вас не уплачен взнос за месяц. Пожалуйсту погасите 
					задолженность, иначе мы будем вынужденны исключить вас из 
					клуба навсегда.
				</p>
				<h3 style="text-align:center;">
					<a href="https://github.com/volitilov">pay this</a>
				</h3>
			</body>
		</html>
	""".format(user)

	msg.attach(MIMEText(html, 'html'))

	smtpObj = smtplib.SMTP_SSL(MAIL_SERVER, MAIL_PORT)
	smtpObj.ehlo()
	smtpObj.login(MAIL_USERNAME, MAIL_PASSWORD)

	smtpObj.sendmail(MAIL_USERNAME, unpaid_members[user], msg.as_string())
	smtpObj.quit()



# TODO: добавить многопоточную отправку